"""
Простые views для веб-интерфейса тестирования СКУД интеграции
"""

from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.core.paginator import Paginator
from datetime import datetime, timedelta, date
from django.db.models import Q, Count
import json
import logging

# Настройка логгера
logger = logging.getLogger(__name__)

from .models import SKUDDevice, SKUDEvent, Employee, WorkDaySummary, WorkSession, Organization, Department, Division
from .skud_device_communication import SKUDDeviceCommunicator, SKUDEventProcessor
from .reports import WorkTimeReportGenerator
from .decorators import (
    require_login, can_view_employee, can_edit_employee, can_manage_skud_devices,
    can_view_reports, can_export_data, can_approve_vacation, can_manage_roles,
    can_create_employee
)
from .permissions import PermissionChecker
from .forms import EmployeeRegistrationForm, EmployeeEditForm, PINFLUpdateForm
from .pinfl_api import pinfl_client, get_employee_by_pinfl_api_view, sync_employee_api_view, create_employee_api_view


@require_login
def admin_test_pinfl_api(request):
    """Административная страница для тестирования PINFL API"""
    # Проверяем, что пользователь является суперпользователем
    if not request.user.is_superuser:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Доступ запрещен. Только администраторы могут использовать эту страницу.")
    
    return render(request, 'employees/test_pinfl_api.html')


@require_login
def dashboard(request):
    """Главная страница с обзором системы"""
    from django.db.models import Count, Q, Sum, Avg
    from datetime import datetime, timedelta
    
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    # Основные метрики
    total_employees = Employee.objects.filter(is_active=True).count()
    active_employees = Employee.objects.filter(is_active=True).count()  # Все активные сотрудники
    
    # Сотрудники в отпуске сегодня
    on_vacation = Employee.objects.filter(
        is_active=True,
        vacations__start_date__lte=today,
        vacations__end_date__gte=today,
        vacations__status__in=['approved', 'taken']
    ).distinct().count()
    
    # Сотрудники в командировке сегодня
    on_business_trip = Employee.objects.filter(
        is_active=True,
        business_trips__start_date__lte=today,
        business_trips__end_date__gte=today,
        business_trips__status__in=['approved', 'in_progress']
    ).distinct().count()
    
    # Сотрудники на больничном сегодня (пока заглушка, так как нет модели больничных)
    on_sick_leave = 0  # TODO: добавить модель больничных
    
    # Сотрудники по полу
    gender_stats = Employee.objects.filter(is_active=True).values('gender').annotate(count=Count('id'))
    male_count = next((item['count'] for item in gender_stats if item['gender'] == 'M'), 0)
    female_count = next((item['count'] for item in gender_stats if item['gender'] == 'F'), 0)
    
    # Получаем параметры фильтрации
    sort_by = request.GET.get('sort', 'time')  # time, name, status
    sort_order = request.GET.get('order', 'asc')  # asc, desc
    
    # Список сотрудников с временем прихода за сегодня
    today_attendance = []
    work_sessions_today = WorkSession.objects.filter(
        date=today,
        employee__is_active=True
    ).select_related('employee').order_by('start_time')
    
    for session in work_sessions_today:
        arrival_time = session.start_time.time()
        is_late = arrival_time > datetime.strptime('09:00', '%H:%M').time()
        
        status = 'on_time' if not is_late else 'late'
        if arrival_time.hour >= 12:  # Если пришел после 12, считаем не пришел
            status = 'absent'
        
        today_attendance.append({
            'employee': session.employee,
            'arrival_time': arrival_time,
            'status': status,
            'is_present': True
        })
    
    # Сотрудники, которые не пришли сегодня
    present_employee_ids = [att['employee'].id for att in today_attendance]
    absent_employees = Employee.objects.filter(
        is_active=True
    ).exclude(id__in=present_employee_ids)
    
    for emp in absent_employees:
        today_attendance.append({
            'employee': emp,
            'arrival_time': None,
            'status': 'absent',
            'is_present': False
        })
    
    # Сортируем по выбранному параметру
    if sort_by == 'name':
        today_attendance.sort(key=lambda x: x['employee'].full_name, reverse=(sort_order == 'desc'))
    elif sort_by == 'status':
        status_order = {'on_time': 0, 'late': 1, 'absent': 2}
        today_attendance.sort(key=lambda x: status_order.get(x['status'], 3), reverse=(sort_order == 'desc'))
    else:  # sort_by == 'time'
        today_attendance.sort(key=lambda x: x['arrival_time'] or datetime.max.time(), reverse=(sort_order == 'desc'))
    
    # Все именинники (сегодня и все остальные)
    all_birthdays = []
    
    # Получаем всех активных сотрудников
    all_employees = Employee.objects.filter(is_active=True).select_related('department', 'division')
    
    for emp in all_employees:
        this_year_birthday = emp.birth_date.replace(year=today.year)
        
        # Если день рождения уже прошел в этом году, берем следующий год
        if this_year_birthday < today:
            next_year_birthday = emp.birth_date.replace(year=today.year + 1)
            days_until = (next_year_birthday - today).days
            birthday_date = next_year_birthday
            is_today = False
        else:
            days_until = (this_year_birthday - today).days
            birthday_date = this_year_birthday
            is_today = (days_until == 0)
        
        all_birthdays.append({
            'employee': emp,
            'days_until': days_until,
            'birthday_date': birthday_date,
            'is_today': is_today
        })
    
    # Сортируем по количеству дней до дня рождения
    all_birthdays.sort(key=lambda x: x['days_until'])
    
    # =============================================================================
    # АНАЛИТИЧЕСКИЕ ДАННЫЕ
    # =============================================================================
    
    # 1. Топ сотрудников по эффективности (за последнюю неделю)
    top_employees = []
    work_sessions_week = WorkSession.objects.filter(
        date__gte=week_ago,
        employee__is_active=True
    ).select_related('employee')
    
    # Группируем по сотрудникам и считаем общее время
    employee_hours = {}
    for session in work_sessions_week:
        emp_id = session.employee.id
        if emp_id not in employee_hours:
            employee_hours[emp_id] = {
                'employee': session.employee,
                'total_hours': 0,
                'sessions_count': 0
            }
        if session.duration_hours:
            employee_hours[emp_id]['total_hours'] += session.duration_hours
        employee_hours[emp_id]['sessions_count'] += 1
    
    # Сортируем по общему времени и берем топ-5
    top_employees = sorted(employee_hours.values(), key=lambda x: x['total_hours'], reverse=True)[:5]
    
    # 2. Проблемы требующие внимания
    problems = []
    
    # Опоздания за неделю
    late_count = 0
    for session in work_sessions_week:
        if session.start_time.time() > datetime.strptime('09:00', '%H:%M').time():
            late_count += 1
    
    if late_count > 0:
        problems.append({
            'type': 'Опоздания',
            'count': late_count,
            'status': 'warning',
            'action': 'Проверить расписание'
        })
    
    # Отсутствие выхода (открытые сессии)
    open_sessions = WorkSession.objects.filter(
        employee__is_active=True,
        end_time__isnull=True,
        date__gte=week_ago
    ).count()
    
    if open_sessions > 0:
        problems.append({
            'type': 'Незакрытые сессии',
            'count': open_sessions,
            'status': 'error',
            'action': 'Закрыть сессии'
        })
    
    # Проблемы с устройствами
    offline_devices = SKUDDevice.objects.filter(
        is_active=True,
        status__in=['inactive', 'error']
    ).count()
    
    if offline_devices > 0:
        problems.append({
            'type': 'Неисправные устройства',
            'count': offline_devices,
            'status': 'error',
            'action': 'Проверить устройства'
        })
    
    # 3. Статистика по отделам
    departments_stats = []
    departments = Employee.objects.filter(is_active=True).values('department__name').annotate(
        total_employees=Count('id')
    ).order_by('-total_employees')[:5]
    
    for dept in departments:
        dept_name = dept['department__name']
        total_emp = dept['total_employees']
        
        # Считаем присутствующих сегодня
        present_today = len([att for att in today_attendance 
                           if att['employee'].department.name == dept_name and att['is_present']])
        
        # Считаем эффективность (часы работы за неделю)
        dept_employees = Employee.objects.filter(
            is_active=True, 
            department__name=dept_name
        )
        total_hours = 0
        for emp in dept_employees:
            emp_sessions = work_sessions_week.filter(employee=emp)
            total_hours += sum(s.duration_hours for s in emp_sessions if s.duration_hours)
        
        efficiency = (total_hours / (total_emp * 40)) * 100 if total_emp > 0 else 0  # 40 часов в неделю
        
        departments_stats.append({
            'name': dept_name,
            'total_employees': total_emp,
            'present_today': present_today,
            'efficiency': round(efficiency, 1)
        })
    
    # 4. Активность СКУД устройств
    devices_activity = []
    devices = SKUDDevice.objects.filter(is_active=True).order_by('name')[:5]
    
    for device in devices:
        # События за сегодня
        today_events = SKUDEvent.objects.filter(
            device=device,
            event_time__date=today
        ).count()
        
        # Последняя активность
        last_event = SKUDEvent.objects.filter(device=device).order_by('-event_time').first()
        last_activity = last_event.event_time if last_event else None
        
        devices_activity.append({
            'device': device,
            'status': device.status,
            'events_today': today_events,
            'last_activity': last_activity
        })
    
    # Данные для графиков
    import json
    
    # Bar chart - топ сотрудники
    top_employees_chart = {
        'labels': [emp['employee'].full_name for emp in top_employees],
        'data': [emp['total_hours'] for emp in top_employees]
    }
    
    # Pie chart - проблемы
    problems_chart = {
        'labels': [p['type'] for p in problems],
        'data': [p['count'] for p in problems]
    }
    
    # Bar chart - отделы
    departments_chart = {
        'labels': [dept['name'] for dept in departments_stats],
        'data': [dept['efficiency'] for dept in departments_stats]
    }
    
    # Line chart - активность устройств за 7 дней
    devices_line_data = []
    for i in range(7):
        date = today - timedelta(days=6-i)
        total_events = SKUDEvent.objects.filter(
            event_time__date=date
        ).count()
        devices_line_data.append(total_events)
    
    devices_chart = {
        'labels': [(today - timedelta(days=6-i)).strftime('%d.%m') for i in range(7)],
        'data': devices_line_data
    }
    
    context = {
        'total_employees': total_employees,
        'active_employees': active_employees,
        'on_vacation': on_vacation,
        'on_business_trip': on_business_trip,
        'on_sick_leave': on_sick_leave,
        'male_count': male_count,
        'female_count': female_count,
        'today_attendance': today_attendance,
        'birthdays': all_birthdays,
        'today': today,
        'sort_by': sort_by,
        'sort_order': sort_order,
        
        # Аналитические данные
        'top_employees': top_employees,
        'problems': problems,
        'departments_stats': departments_stats,
        'devices_activity': devices_activity,
        
        # Данные для графиков (JSON сериализованные)
        'top_employees_chart': json.dumps(top_employees_chart),
        'problems_chart': json.dumps(problems_chart),
        'departments_chart': json.dumps(departments_chart),
        'devices_chart': json.dumps(devices_chart),
    }
    
    return render(request, 'employees/dashboard.html', context)


@require_login
def analytics_data(request):
    """AJAX endpoint для получения данных аналитики с фильтрами"""
    from django.http import JsonResponse
    from django.db.models import Count, Q, Sum, Avg
    from datetime import datetime, timedelta
    import json
    
    # Получаем параметры фильтрации
    chart_type = request.GET.get('chart_type')
    period = int(request.GET.get('period', 7))
    limit = int(request.GET.get('limit', 5))
    filter_type = request.GET.get('filter_type', 'all')
    metric = request.GET.get('metric', 'efficiency')
    
    today = timezone.now().date()
    start_date = today - timedelta(days=period-1)
    
    try:
        if chart_type == 'top_employees':
            # Топ сотрудников
            work_sessions = WorkSession.objects.filter(
                date__gte=start_date,
                employee__is_active=True
            ).select_related('employee')
            
            employee_hours = {}
            for session in work_sessions:
                emp_id = session.employee.id
                if emp_id not in employee_hours:
                    employee_hours[emp_id] = {
                        'employee': session.employee,
                        'total_hours': 0,
                        'sessions_count': 0
                    }
                if session.duration_hours:
                    employee_hours[emp_id]['total_hours'] += session.duration_hours
                employee_hours[emp_id]['sessions_count'] += 1
            
            top_employees = sorted(employee_hours.values(), key=lambda x: x['total_hours'], reverse=True)[:limit]
            
            data = {
                'labels': [emp['employee'].full_name for emp in top_employees],
                'data': [emp['total_hours'] for emp in top_employees],
                'table_data': [
                    {
                        'name': emp['employee'].full_name,
                        'department': emp['employee'].department.name,
                        'hours': emp['total_hours'],
                        'sessions': emp['sessions_count']
                    } for emp in top_employees
                ]
            }
            
        elif chart_type == 'problems':
            # Проблемы
            problems = []
            
            # Опоздания
            if filter_type in ['all', 'late']:
                work_sessions = WorkSession.objects.filter(
                    date__gte=start_date,
                    employee__is_active=True
                )
                late_count = 0
                for session in work_sessions:
                    if session.start_time.time() > datetime.strptime('09:00', '%H:%M').time():
                        late_count += 1
                
                if late_count > 0:
                    problems.append({
                        'type': 'Опоздания',
                        'count': late_count,
                        'status': 'warning'
                    })
            
            # Незакрытые сессии
            if filter_type in ['all', 'sessions']:
                open_sessions = WorkSession.objects.filter(
                    employee__is_active=True,
                    end_time__isnull=True,
                    date__gte=start_date
                ).count()
                
                if open_sessions > 0:
                    problems.append({
                        'type': 'Незакрытые сессии',
                        'count': open_sessions,
                        'status': 'error'
                    })
            
            # Проблемы с устройствами
            if filter_type in ['all', 'devices']:
                offline_devices = SKUDDevice.objects.filter(
                    is_active=True,
                    status__in=['inactive', 'error']
                ).count()
                
                if offline_devices > 0:
                    problems.append({
                        'type': 'Неисправные устройства',
                        'count': offline_devices,
                        'status': 'error'
                    })
            
            data = {
                'labels': [p['type'] for p in problems],
                'data': [p['count'] for p in problems],
                'table_data': problems
            }
            
        elif chart_type == 'departments':
            # Отделы
            departments = Employee.objects.filter(is_active=True).values('department__name').annotate(
                total_employees=Count('id')
            ).order_by('-total_employees')[:limit]
            
            departments_stats = []
            work_sessions = WorkSession.objects.filter(
                date__gte=start_date,
                employee__is_active=True
            )
            
            for dept in departments:
                dept_name = dept['department__name']
                total_emp = dept['total_employees']
                
                # Считаем присутствующих сегодня
                present_today = Employee.objects.filter(
                    is_active=True,
                    department__name=dept_name
                ).count()
                
                # Считаем метрику
                dept_employees = Employee.objects.filter(
                    is_active=True, 
                    department__name=dept_name
                )
                
                if metric == 'efficiency':
                    total_hours = 0
                    for emp in dept_employees:
                        emp_sessions = work_sessions.filter(employee=emp)
                        total_hours += sum(s.duration_hours for s in emp_sessions if s.duration_hours)
                    value = (total_hours / (total_emp * 40)) * 100 if total_emp > 0 else 0
                elif metric == 'attendance':
                    present_days = 0
                    total_days = 0
                    for emp in dept_employees:
                        emp_sessions = work_sessions.filter(employee=emp)
                        present_days += emp_sessions.count()
                        total_days += period
                    value = (present_days / total_days) * 100 if total_days > 0 else 0
                else:  # hours
                    total_hours = 0
                    for emp in dept_employees:
                        emp_sessions = work_sessions.filter(employee=emp)
                        total_hours += sum(s.duration_hours for s in emp_sessions if s.duration_hours)
                    value = total_hours
                
                departments_stats.append({
                    'name': dept_name,
                    'total_employees': total_emp,
                    'present_today': present_today,
                    'value': round(value, 1)
                })
            
            data = {
                'labels': [dept['name'] for dept in departments_stats],
                'data': [dept['value'] for dept in departments_stats],
                'table_data': departments_stats
            }
            
        elif chart_type == 'devices':
            # Устройства
            devices_query = SKUDDevice.objects.filter(is_active=True)
            if filter_type == 'active':
                devices_query = devices_query.filter(status='active')
            elif filter_type == 'inactive':
                devices_query = devices_query.filter(status__in=['inactive', 'error'])
            
            devices = devices_query.order_by('name')[:limit]
            
            devices_activity = []
            for device in devices:
                # События за период
                period_events = SKUDEvent.objects.filter(
                    device=device,
                    event_time__date__gte=start_date
                ).count()
                
                # Последняя активность
                last_event = SKUDEvent.objects.filter(device=device).order_by('-event_time').first()
                last_activity = last_event.event_time if last_event else None
                
                devices_activity.append({
                    'device_name': device.name,
                    'location': device.location,
                    'status': device.status,
                    'events_count': period_events,
                    'last_activity': last_activity
                })
            
            # Данные для линейного графика
            line_data = []
            for i in range(period):
                date = today - timedelta(days=period-1-i)
                total_events = SKUDEvent.objects.filter(
                    event_time__date=date
                ).count()
                line_data.append(total_events)
            
            data = {
                'labels': [(today - timedelta(days=period-1-i)).strftime('%d.%m') for i in range(period)],
                'data': line_data,
                'table_data': devices_activity
            }
        
        else:
            return JsonResponse({'error': 'Invalid chart type'}, status=400)
        
        return JsonResponse({
            'success': True,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_login
@can_manage_skud_devices
def devices_list(request):
    """Список СКУД устройств"""
    # Фильтр по статусу активности
    show_inactive = request.GET.get('show_inactive', 'false').lower() == 'true'
    
    # Оптимизированный запрос с select_related для будущих расширений
    if show_inactive:
        devices = SKUDDevice.objects.all().order_by('-is_active', 'name')
    else:
        devices = SKUDDevice.objects.filter(is_active=True).order_by('name')
    
    # Пагинация
    paginator = Paginator(devices, 12)  # Увеличили до 12 для лучшего отображения
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'devices': page_obj,
        'show_inactive': show_inactive,
    }
    
    return render(request, 'employees/devices_list.html', context)


@require_login
@can_manage_skud_devices
def device_detail(request, device_id):
    """Детальная информация об устройстве"""
    try:
        device = SKUDDevice.objects.get(id=device_id)
        
        # Последние события устройства
        events = SKUDEvent.objects.filter(device=device).order_by('-event_time')[:20]
        
        # Статистика за последние 7 дней
        week_ago = timezone.now() - timedelta(days=7)
        week_events = SKUDEvent.objects.filter(
            device=device,
            event_time__gte=week_ago
        )
        
        context = {
            'device': device,
            'events': events,
            'week_events_count': week_events.count(),
            'entry_count': week_events.filter(event_type='entry').count(),
            'exit_count': week_events.filter(event_type='exit').count(),
        }
        
        return render(request, 'employees/device_detail.html', context)
        
    except SKUDDevice.DoesNotExist:
        messages.error(request, 'Устройство не найдено')
        return redirect('devices_list')


@require_login
def events_list(request):
    """Список событий СКУД"""
    # Оптимизированный запрос с select_related
    events = SKUDEvent.objects.select_related('device', 'employee').order_by('-event_time')
    
    # Фильтры
    device_id = request.GET.get('device')
    event_type = request.GET.get('type')
    hours = int(request.GET.get('hours', 24))
    
    if device_id:
        events = events.filter(device_id=device_id)
    if event_type:
        events = events.filter(event_type=event_type)
    
    # Фильтр по времени
    since_time = timezone.now() - timedelta(hours=hours)
    events = events.filter(event_time__gte=since_time)
    
    # Пагинация
    paginator = Paginator(events, 25)  # Увеличили для лучшей производительности
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Оптимизированные данные для фильтров - только активные устройства
    devices = SKUDDevice.objects.filter(is_active=True).order_by('name')
    
    context = {
        'page_obj': page_obj,
        'events': page_obj,
        'devices': devices,
        'current_device': device_id,
        'current_type': event_type,
        'current_hours': hours,
    }
    
    return render(request, 'employees/events_list.html', context)


def test_device(request, device_id):
    """Тестирование устройства"""
    try:
        device = SKUDDevice.objects.get(id=device_id)
        communicator = SKUDDeviceCommunicator()
        
        is_online, message = communicator.test_device_connection(device)
        
        return JsonResponse({
            'success': is_online,
            'message': message,
            'device_name': device.name
        })
        
    except SKUDDevice.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Устройство не найдено'
        })


@require_login
@can_manage_skud_devices
def add_device(request):
    """Добавление нового устройства"""
    if request.method == 'POST':
        name = request.POST.get('name')
        ip_address = request.POST.get('ip_address')
        port = int(request.POST.get('port', 80))
        serial_number = request.POST.get('serial_number')
        device_type = request.POST.get('device_type')
        location = request.POST.get('location', '')
        description = request.POST.get('description', '')
        
        try:
            device = SKUDDevice.objects.create(
                name=name,
                ip_address=ip_address,
                port=port,
                serial_number=serial_number,
                device_type=device_type,
                location=location,
                description=description
            )
            
            messages.success(request, f'Устройство "{device.name}" успешно добавлено')
            return redirect('device_detail', device_id=device.id)
            
        except Exception as e:
            messages.error(request, f'Ошибка при добавлении устройства: {str(e)}')
    
    return render(request, 'employees/add_device.html')


def send_test_event(request):
    """Отправка тестового события"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Получаем IP адрес клиента
            client_ip = request.META.get('REMOTE_ADDR', '127.0.0.1')
            
            # Обрабатываем событие
            communicator = SKUDDeviceCommunicator()
            skud_event = communicator.process_device_event(client_ip, data)
            
            return JsonResponse({
                'success': True,
                'message': 'Событие успешно обработано',
                'event_id': str(skud_event.id),
                'employee_found': skud_event.employee is not None
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Ошибка: {str(e)}'
            })
    
    return render(request, 'employees/send_test_event.html')


def api_status(request):
    """API статус системы"""
    from .cache_utils import SKUDCache
    
    try:
        # Проверяем кэш
        cached_status = SKUDCache.get_api_status()
        
        if cached_status:
            return JsonResponse(cached_status)
        
        # Если кэш пустой, генерируем данные (БЕЗ проверки устройств!)
        from django.db.models import Count, Q
        
        device_stats = SKUDDevice.objects.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(is_active=True))
        )
        
        event_stats = SKUDEvent.objects.aggregate(
            total=Count('id'),
            today=Count('id', filter=Q(event_time__date=timezone.now().date()))
        )
        
        status_data = {
            'status': 'success',
            'total_devices': device_stats['total'],
            'active_devices': device_stats['active'],
            'total_events': event_stats['total'],
            'today_events': event_stats['today'],
            'server_time': timezone.now().isoformat(),
            'devices': {}  # Пустой объект устройств - проверка по запросу
        }
        
        # Кэшируем результат
        SKUDCache.set_api_status(status_data)
        
        return JsonResponse(status_data)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e),
            'server_time': timezone.now().isoformat()
        }, status=500)


def check_devices_health(request):
    """AJAX endpoint для проверки статуса устройств"""
    if request.method == 'POST':
        communicator = SKUDDeviceCommunicator()
        results = communicator.check_all_devices_health()
        
        return JsonResponse({
            'status': 'success',
            'devices': results,
            'timestamp': timezone.now().isoformat()
        })
    
    return JsonResponse({
        'status': 'error',
        'message': 'Only POST requests allowed'
    }, status=405)




@require_login
def employees_list(request):
    """Список сотрудников с поиском и фильтрацией"""
    from django.db.models import Q
    
    # Базовый queryset - фильтруем по правам доступа
    if request.user.is_superuser:
        employees = Employee.objects.filter(is_active=True).select_related('department', 'division')
    else:
        # Получаем доступных сотрудников через систему ролей
        accessible_employees = PermissionChecker.get_accessible_employees(request.user)
        employees = accessible_employees.select_related('department', 'division')
    
    # Поиск по ФИО
    search_query = request.GET.get('search', '').strip()
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(middle_name__icontains=search_query)
        )
    
    # Фильтр по отделу
    department_id = request.GET.get('department')
    if department_id:
        employees = employees.filter(department_id=department_id)
    
    # Фильтр по должности
    position = request.GET.get('position')
    if position:
        employees = employees.filter(position=position)
    
    # Фильтр по статусу
    status = request.GET.get('status')
    if status == 'active':
        employees = employees.filter(is_active=True)
    elif status == 'inactive':
        employees = employees.filter(is_active=False)
    
    # Сортировка
    sort_by = request.GET.get('sort', 'last_name')
    sort_order = request.GET.get('order', 'asc')
    
    if sort_order == 'desc':
        sort_by = f'-{sort_by}'
    
    employees = employees.order_by(sort_by)
    
    # Количество записей на странице
    per_page = int(request.GET.get('per_page', 20))
    per_page = min(per_page, 100)  # Максимум 100 записей
    
    # Пагинация
    paginator = Paginator(employees, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Данные для фильтров
    departments = Employee.objects.filter(
        is_active=True,
        department__isnull=False
    ).values_list('department__id', 'department__name').distinct().order_by('department__name')
    
    positions = Employee.POSITION_CHOICES
    
    context = {
        'page_obj': page_obj,
        'employees': page_obj,
        'departments': departments,
        'positions': positions,
        'search_query': search_query,
        'current_filters': {
            'department': department_id,
            'position': position,
            'status': status,
        }
    }
    
    return render(request, 'employees/employees_list.html', context)


@require_login
@can_view_employee
def employee_events(request, employee_id):
    """События конкретного сотрудника"""
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # События за последние 30 дней
        month_ago = timezone.now() - timedelta(days=30)
        events = SKUDEvent.objects.filter(
            employee=employee,
            event_time__gte=month_ago
        ).order_by('-event_time')
        
        context = {
            'employee': employee,
            'events': events,
        }
        
        return render(request, 'employees/employee_events.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Сотрудник не найден')
        return redirect('employees_list')


@require_login
@can_manage_skud_devices
def delete_device(request, device_id):
    """Удаление СКУД устройства"""
    try:
        device = SKUDDevice.objects.get(id=device_id)
        
        if request.method == 'POST':
            # Проверяем, есть ли связанные события
            events_count = SKUDEvent.objects.filter(device=device).count()
            
            if events_count > 0:
                messages.warning(
                    request, 
                    f'Нельзя удалить устройство "{device.name}" - у него есть {events_count} связанных событий. '
                    'Сначала удалите все события или деактивируйте устройство.'
                )
                return redirect('device_detail', device_id=device.id)
            
            # Удаляем устройство
            device_name = device.name
            device.delete()
            
            messages.success(request, f'Устройство "{device_name}" успешно удалено')
            return redirect('devices_list')
        
        # GET запрос - показываем страницу подтверждения
        context = {
            'device': device,
        }
        return render(request, 'employees/delete_device.html', context)
        
    except SKUDDevice.DoesNotExist:
        messages.error(request, 'Устройство не найдено')
        return redirect('devices_list')


@require_login
@can_manage_skud_devices
def deactivate_device(request, device_id):
    """Деактивация СКУД устройства (мягкое удаление)"""
    try:
        device = SKUDDevice.objects.get(id=device_id)
        
        if request.method == 'POST':
            # Деактивируем устройство вместо удаления
            device.is_active = False
            device.status = 'inactive'
            device.save()
            
            messages.success(request, f'Устройство "{device.name}" деактивировано')
            return redirect('device_detail', device_id=device.id)
        
        # GET запрос - показываем страницу подтверждения
        context = {
            'device': device,
        }
        return render(request, 'employees/deactivate_device.html', context)
        
    except SKUDDevice.DoesNotExist:
        messages.error(request, 'Устройство не найдено')
        return redirect('devices_list')


@require_login
@can_manage_skud_devices
def activate_device(request, device_id):
    """Активация СКУД устройства"""
    try:
        device = SKUDDevice.objects.get(id=device_id)
        
        if request.method == 'POST':
            # Активируем устройство
            device.is_active = True
            device.status = 'active'
            device.save()
            
            messages.success(request, f'Устройство "{device.name}" активировано')
            return redirect('device_detail', device_id=device.id)
        
        # GET запрос - показываем страницу подтверждения
        context = {
            'device': device,
        }
        return render(request, 'employees/activate_device.html', context)
        
    except SKUDDevice.DoesNotExist:
        messages.error(request, 'Устройство не найдено')
        return redirect('devices_list')


def login_view(request):
    """Страница входа в систему"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f'Добро пожаловать, {user.first_name or user.username}!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Неверные учетные данные')
    
    return render(request, 'employees/login.html')


def logout_view(request):
    """Выход из системы"""
    logout(request)
    messages.info(request, 'Вы вышли из системы')
    return redirect('login_view')


@require_login
def profile_view(request):
    """Профиль пользователя"""
    return render(request, 'employees/profile.html')


@require_login
@can_view_reports
def reports_dashboard(request):
    """Дашборд отчетов"""
    from .reports import WorkTimeReportGenerator
    from datetime import datetime, date
    
    # Получаем текущую дату
    now = datetime.now()
    current_year = now.year
    current_month = now.month
    
    # Получаем доступных сотрудников для отчетов
    if request.user.is_superuser:
        employees = Employee.objects.filter(is_active=True).select_related('department', 'division')
    else:
        employees = PermissionChecker.get_accessible_employees(request.user)
    
    # Статистика для дашборда
    total_employees = employees.count()
    active_employees = employees.filter(is_active=True).count()
    
    # Получаем статистику по департаментам
    departments_stats = employees.values('department__name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    context = {
        'current_year': current_year,
        'current_month': current_month,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'departments_stats': departments_stats,
        'employees': employees[:10],  # Последние 10 сотрудников
    }
    
    return render(request, 'employees/reports_dashboard.html', context)


@require_login
def work_time_summaries(request):
    """Сводки рабочего времени"""
    employee_id = request.GET.get('employee_id')
    
    # Если указан employee_id, показываем сводки для конкретного сотрудника
    if employee_id:
        try:
            employee = Employee.objects.get(id=employee_id)
            # Проверяем права доступа
            if not PermissionChecker.can_view_employee(request.user, employee):
                messages.error(request, 'У вас нет прав для просмотра данных этого сотрудника')
                return redirect('profile_view')
        except Employee.DoesNotExist:
            messages.error(request, 'Сотрудник не найден')
            return redirect('profile_view')
    else:
        # Если не указан, показываем для текущего пользователя
        if hasattr(request.user, 'employee_profile') and request.user.employee_profile:
            employee = request.user.employee_profile
        else:
            messages.error(request, 'У вас нет профиля сотрудника')
            return redirect('profile_view')
    
    # Получаем сводки за последние 30 дней
    from datetime import timedelta
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    summaries = WorkDaySummary.objects.filter(
        employee=employee,
        date__range=[start_date, end_date]
    ).order_by('-date')
    
    context = {
        'employee': employee,
        'summaries': summaries,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'employees/work_time_summaries.html', context)


@require_login
def work_sessions(request):
    """Сессии рабочего времени"""
    employee_id = request.GET.get('employee_id')
    
    # Если указан employee_id, показываем сессии для конкретного сотрудника
    if employee_id:
        try:
            employee = Employee.objects.get(id=employee_id)
            # Проверяем права доступа
            if not PermissionChecker.can_view_employee(request.user, employee):
                messages.error(request, 'У вас нет прав для просмотра данных этого сотрудника')
                return redirect('profile_view')
        except Employee.DoesNotExist:
            messages.error(request, 'Сотрудник не найден')
            return redirect('profile_view')
    else:
        # Если не указан, показываем для текущего пользователя
        if hasattr(request.user, 'employee_profile') and request.user.employee_profile:
            employee = request.user.employee_profile
        else:
            messages.error(request, 'У вас нет профиля сотрудника')
            return redirect('profile_view')
    
    # Получаем сессии за последние 7 дней
    from datetime import timedelta
    end_date = timezone.now()
    start_date = end_date - timedelta(days=7)
    
    sessions = WorkSession.objects.filter(
        employee=employee,
        start_time__range=[start_date, end_date]
    ).order_by('-start_time')
    
    context = {
        'employee': employee,
        'sessions': sessions,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'employees/work_sessions.html', context)


@require_login
def employee_report(request):
    """Отчет по сотруднику"""
    # Показываем отчет для текущего пользователя
    if hasattr(request.user, 'employee_profile') and request.user.employee_profile:
        employee = request.user.employee_profile
    else:
        messages.error(request, 'У вас нет профиля сотрудника')
        return redirect('profile_view')
    
    # Получаем данные за текущий месяц
    from datetime import datetime
    now = datetime.now()
    start_date = now.replace(day=1).date()
    
    # Получаем сводки за месяц
    summaries = WorkDaySummary.objects.filter(
        employee=employee,
        date__gte=start_date
    ).order_by('-date')
    
    # Подсчитываем статистику
    total_days = summaries.count()
    total_hours = sum(s.total_work_hours for s in summaries if s.total_work_hours)
    avg_hours = total_hours / total_days if total_days > 0 else 0
    
    context = {
        'employee': employee,
        'summaries': summaries,
        'start_date': start_date,
        'total_days': total_days,
        'total_hours': total_hours,
        'avg_hours': avg_hours,
    }
    
    return render(request, 'employees/employee_report.html', context)


@require_login
@can_view_reports
def attendance_control(request):
    """Страница контроля прибытия и отбытия"""
    from django.db.models import Q, Count, Sum, Avg
    from datetime import datetime, timedelta, date
    import calendar
    
    # Получаем параметры фильтрации
    view_type = request.GET.get('view', 'daily')  # daily или monthly
    search_query = request.GET.get('search', '').strip()
    department_id = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    
    # Определяем диапазон дат для фильтрации
    today = timezone.now().date()
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = today
    else:
        date_from = today
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            date_to = today
    else:
        date_to = today
    
    # Убеждаемся, что date_from не больше date_to
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    
    # Базовый queryset сотрудников - фильтруем по правам доступа
    if request.user.is_superuser:
        employees = Employee.objects.filter(is_active=True).select_related('department', 'division')
    else:
        # Получаем доступных сотрудников через систему ролей
        accessible_employees = PermissionChecker.get_accessible_employees(request.user)
        employees = accessible_employees.select_related('department', 'division')
    
    # Поиск по имени
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(middle_name__icontains=search_query)
        )
    
    # Фильтр по отделу
    if department_id:
        employees = employees.filter(department_id=department_id)
    
    # Получаем данные о посещаемости
    attendance_data = []
    
    if view_type == 'daily':
        # Дневной вид - показываем данные за диапазон дней
        for employee in employees:
            # Получаем рабочие сессии за диапазон дней
            work_sessions = WorkSession.objects.filter(
                employee=employee,
                date__range=[date_from, date_to]
            ).order_by('date', 'start_time')
            
            # Получаем сводки за диапазон дней
            day_summaries = WorkDaySummary.objects.filter(
                employee=employee,
                date__range=[date_from, date_to]
            ).order_by('date')
            
            # Рассчитываем статистику за весь диапазон
            total_hours = 0
            entry_count = 0
            arrival_time = None
            departure_time = None
            
            if work_sessions.exists():
                first_session = work_sessions.first()
                last_session = work_sessions.last()
                
                arrival_time = first_session.start_time.time()
                if last_session.end_time:
                    departure_time = last_session.end_time.time()
                
                entry_count = work_sessions.count()
                total_hours = sum(session.duration_hours for session in work_sessions if session.duration_hours)
            
            # Определяем статус на основе сводок за период
            status = 'absent'
            if day_summaries.exists():
                # Если есть сводки, определяем общий статус
                present_days = day_summaries.filter(status='present').count()
                total_days = day_summaries.count()
                
                if present_days == total_days:
                    status = 'present'
                elif present_days > 0:
                    status = 'partial'
                else:
                    status = 'absent'
            elif work_sessions.exists():
                status = 'present'
            
            # Применяем фильтр по статусу
            if status_filter and status != status_filter:
                continue
            
            attendance_data.append({
                'employee': employee,
                'arrival_time': arrival_time,
                'departure_time': departure_time,
                'total_hours': total_hours,
                'entry_count': entry_count,
                'status': status,
                'work_sessions': work_sessions,
                'day_summaries': day_summaries,
            })
    
    else:  # monthly
        # Месячный вид - показываем статистику за диапазон дат
        for employee in employees:
            # Получаем сводки за диапазон дат
            month_summaries = WorkDaySummary.objects.filter(
                employee=employee,
                date__range=[date_from, date_to]
            )
            
            # Рассчитываем общую статистику
            total_work_hours = sum(s.total_hours for s in month_summaries)
            total_expected_hours = sum(s.expected_hours for s in month_summaries)
            # Изменено: считаем количество уникальных дней вместо количества входов
            total_entry_count = month_summaries.filter(sessions_count__gt=0).count()
            
            # Статус на основе количества присутствий
            present_days = month_summaries.filter(status='present').count()
            total_work_days = month_summaries.count()
            
            if total_work_days == 0:
                status = 'no_data'
            elif present_days == total_work_days:
                status = 'excellent'
            elif present_days >= total_work_days * 0.8:
                status = 'good'
            elif present_days >= total_work_days * 0.6:
                status = 'average'
            else:
                status = 'poor'
            
            # Применяем фильтр по статусу
            if status_filter and status != status_filter:
                continue
            
            attendance_data.append({
                'employee': employee,
                'total_work_hours': total_work_hours,
                'total_expected_hours': total_expected_hours,
                'total_entry_count': total_entry_count,
                'present_days': present_days,
                'total_work_days': total_work_days,
                'status': status,
                'month_summaries': month_summaries,
            })
    
    # Сортируем данные
    sort_by = request.GET.get('sort', 'name')
    sort_order = request.GET.get('order', 'asc')
    
    if sort_by == 'name':
        attendance_data.sort(key=lambda x: x['employee'].full_name, reverse=(sort_order == 'desc'))
    elif sort_by == 'hours':
        attendance_data.sort(key=lambda x: x.get('total_hours', 0) or x.get('total_work_hours', 0), reverse=(sort_order == 'desc'))
    elif sort_by == 'entries':
        attendance_data.sort(key=lambda x: x.get('entry_count', 0) or x.get('total_entry_count', 0), reverse=(sort_order == 'desc'))
    
    # Пагинация
    per_page = int(request.GET.get('per_page', 20))
    paginator = Paginator(attendance_data, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Данные для фильтров
    departments = Employee.objects.filter(
        is_active=True,
        department__isnull=False
    ).values_list('department__id', 'department__name').distinct().order_by('department__name')
    
    # Статистика для заголовка
    total_employees = len(attendance_data)
    present_count = len([d for d in attendance_data if d['status'] in ['present', 'partial', 'excellent', 'good']])
    
    context = {
        'page_obj': page_obj,
        'attendance_data': page_obj,
        'view_type': view_type,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'department_id': department_id,
        'status_filter': status_filter,
        'departments': departments,
        'total_employees': total_employees,
        'present_count': present_count,
        'sort_by': sort_by,
        'sort_order': sort_order,
        'per_page': per_page,
    }
    
    return render(request, 'employees/attendance_control.html', context)


@require_login
@can_export_data
def export_attendance_excel(request):
    """Экспорт данных о посещаемости в Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from datetime import datetime, timedelta, date
    from django.db.models import Q
    
    # Получаем параметры фильтрации (те же, что и в attendance_control)
    view_type = request.GET.get('view', 'daily')
    search_query = request.GET.get('search', '').strip()
    department_id = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    
    # Определяем диапазон дат для фильтрации
    today = timezone.now().date()
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            date_from = today
    else:
        date_from = today
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            date_to = today
    else:
        date_to = today
    
    # Убеждаемся, что date_from не больше date_to
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Посещаемость_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    if date_from == date_to:
        title = f"Отчет по посещаемости - {date_from.strftime('%d.%m.%Y')}"
    else:
        title = f"Отчет по посещаемости с {date_from.strftime('%d.%m.%Y')} по {date_to.strftime('%d.%m.%Y')}"
    
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    # Заголовки колонок
    if view_type == 'daily':
        headers = ['Сотрудник', 'Должность', 'Отдел', 'Время прихода', 'Время ухода', 'Часов работы', 'Статус']
    else:
        headers = ['Сотрудник', 'Должность', 'Отдел', 'Часов работы', 'Ожидаемо часов', 'Количество дней', 'Статус']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Получаем данные (используем ту же логику, что и в attendance_control)
    # Фильтруем по правам доступа
    if request.user.is_superuser:
        employees = Employee.objects.filter(is_active=True).select_related('department', 'division')
    else:
        # Получаем доступных сотрудников через систему ролей
        accessible_employees = PermissionChecker.get_accessible_employees(request.user)
        employees = accessible_employees.select_related('department', 'division')
    
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(middle_name__icontains=search_query)
        )
    
    if department_id:
        employees = employees.filter(department_id=department_id)
    
    row = 4
    for employee in employees:
        if view_type == 'daily':
            # Дневной вид - данные за диапазон дат
            work_sessions = WorkSession.objects.filter(
                employee=employee,
                date__range=[date_from, date_to]
            ).order_by('date', 'start_time')
            
            day_summaries = WorkDaySummary.objects.filter(
                employee=employee,
                date__range=[date_from, date_to]
            ).order_by('date')
            
            arrival_time = None
            departure_time = None
            total_hours = 0
            
            if work_sessions.exists():
                first_session = work_sessions.first()
                last_session = work_sessions.last()
                
                arrival_time = first_session.start_time.time()
                if last_session.end_time:
                    departure_time = last_session.end_time.time()
                
                total_hours = sum(session.duration_hours for session in work_sessions if session.duration_hours)
            
            # Определяем статус на основе сводок за период
            status = 'Отсутствовал'
            if day_summaries.exists():
                present_days = day_summaries.filter(status='present').count()
                total_days = day_summaries.count()
                
                if present_days == total_days:
                    status = 'Присутствовал'
                elif present_days > 0:
                    status = 'Частично присутствовал'
                else:
                    status = 'Отсутствовал'
            elif work_sessions.exists():
                status = 'Присутствовал'
            
            # Применяем фильтр по статусу
            if status_filter and status.lower() != status_filter:
                continue
            
            data = [
                employee.full_name,
                employee.get_position_display(),
                employee.department.name,
                arrival_time.strftime('%H:%M') if arrival_time else '—',
                departure_time.strftime('%H:%M') if departure_time else '—',
                f"{total_hours:.1f}" if total_hours > 0 else '—',
                status
            ]
        else:
            # Месячный вид - данные за диапазон дат
            month_summaries = WorkDaySummary.objects.filter(
                employee=employee,
                date__range=[date_from, date_to]
            )
            
            total_work_hours = sum(s.total_hours for s in month_summaries)
            total_expected_hours = sum(s.expected_hours for s in month_summaries)
            # Изменено: считаем количество уникальных дней вместо количества входов
            total_entry_count = month_summaries.filter(sessions_count__gt=0).count()
            
            present_days = month_summaries.filter(status='present').count()
            total_work_days = month_summaries.count()
            
            if total_work_days == 0:
                status = 'Нет данных'
            elif present_days == total_work_days:
                status = 'Отлично'
            elif present_days >= total_work_days * 0.8:
                status = 'Хорошо'
            elif present_days >= total_work_days * 0.6:
                status = 'Удовлетворительно'
            else:
                status = 'Плохо'
            
            # Применяем фильтр по статусу
            if status_filter and status.lower() != status_filter:
                continue
            
            data = [
                employee.full_name,
                employee.get_position_display(),
                employee.department.name,
                f"{total_work_hours:.1f}",
                f"{total_expected_hours:.1f}",
                str(total_entry_count),
                status
            ]
        
        # Записываем данные в Excel
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
    
    # Автоподбор ширины колонок
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Создаем HTTP ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attendance_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}_{view_type}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@require_login
@can_create_employee
def create_employee(request):
    """Создание нового сотрудника"""
    if request.method == 'POST':
        form = EmployeeRegistrationForm(request.POST)
        if form.is_valid():
            try:
                employee = form.save()
                messages.success(request, f'Сотрудник {employee.full_name} успешно создан')
                
                # Автоматически синхронизируем PINFL с SKUD API
                if employee.pinfl:
                    sync_result = pinfl_client.sync_employee_pinfl(
                        employee=employee,
                        pinfl=employee.pinfl,
                        date=date.today()
                    )
                    
                    if sync_result['success']:
                        messages.success(request, 'PINFL успешно синхронизирован с SKUD')
                    else:
                        messages.warning(request, f'PINFL не синхронизирован с SKUD: {sync_result.get("error_details", "Неизвестная ошибка")}')
                
                return redirect('employee_detail', employee_id=employee.id)
            except Exception as e:
                messages.error(request, f'Ошибка при создании сотрудника: {str(e)}')
    else:
        form = EmployeeRegistrationForm()
    
    # Получаем данные для выпадающих списков
    organizations = Organization.objects.all()
    departments = Department.objects.all()
    divisions = Division.objects.all()
    
    context = {
        'form': form,
        'organizations': organizations,
        'departments': departments,
        'divisions': divisions,
    }
    
    return render(request, 'employees/create_employee.html', context)


@require_login
def employee_detail(request, employee_id):
    """Детальная информация о сотруднике (только для администраторов)"""
    # Проверяем, что пользователь - администратор
    if not request.user.is_superuser:
        messages.error(request, 'Доступ к детальной информации о сотрудниках разрешен только администраторам')
        return redirect('employees_list')
    
    try:
        employee = Employee.objects.select_related('user', 'organization', 'department', 'division').get(id=employee_id)
        
        # Получаем последние события сотрудника
        recent_events = SKUDEvent.objects.filter(employee=employee).order_by('-event_time')[:10]
        
        # Получаем сводки за последние 7 дней
        week_ago = timezone.now().date() - timedelta(days=7)
        recent_summaries = WorkDaySummary.objects.filter(
            employee=employee,
            date__gte=week_ago
        ).order_by('-date')
        
        context = {
            'employee': employee,
            'recent_events': recent_events,
            'recent_summaries': recent_summaries,
        }
        
        return render(request, 'employees/employee_detail.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Сотрудник не найден')
        return redirect('employees_list')


@require_login
@can_edit_employee
def edit_employee(request, employee_id):
    """Редактирование сотрудника"""
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # Проверяем права доступа
        if not PermissionChecker.can_edit_employee(request.user, employee):
            messages.error(request, 'У вас нет прав для редактирования данных этого сотрудника')
            return redirect('employee_detail', employee_id=employee.id)
        
        if request.method == 'POST':
            form = EmployeeEditForm(request.POST, instance=employee, employee=employee)
            if form.is_valid():
                try:
                    form.save()
                    messages.success(request, f'Данные сотрудника {employee.full_name} успешно обновлены')
                    return redirect('employee_detail', employee_id=employee.id)
                except Exception as e:
                    messages.error(request, f'Ошибка при обновлении данных: {str(e)}')
        else:
            form = EmployeeEditForm(instance=employee, employee=employee)
        
        # Получаем данные для выпадающих списков
        organizations = Organization.objects.all()
        departments = Department.objects.all()
        divisions = Division.objects.all()
        
        context = {
            'form': form,
            'employee': employee,
            'organizations': organizations,
            'departments': departments,
            'divisions': divisions,
        }
        
        return render(request, 'employees/edit_employee.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Сотрудник не найден')
        return redirect('employees_list')


@require_login
@can_edit_employee
def update_pinfl(request, employee_id):
    """Обновление PINFL сотрудника"""
    try:
        employee = Employee.objects.get(id=employee_id)
        
        # Проверяем права доступа
        if not PermissionChecker.can_edit_employee(request.user, employee):
            messages.error(request, 'У вас нет прав для редактирования данных этого сотрудника')
            return redirect('employee_detail', employee_id=employee.id)
        
        if request.method == 'POST':
            form = PINFLUpdateForm(request.POST, employee=employee)
            if form.is_valid():
                try:
                    old_pinfl = employee.pinfl
                    new_pinfl = form.cleaned_data['pinfl']
                    
                    # Обновляем PINFL
                    employee.pinfl = new_pinfl
                    employee.save(update_fields=['pinfl'])
                    
                    # Синхронизируем с SKUD API
                    sync_result = pinfl_client.sync_employee_pinfl(
                        employee=employee,
                        pinfl=new_pinfl,
                        date=date.today()
                    )
                    
                    if sync_result['success']:
                        messages.success(request, f'PINFL успешно обновлен и синхронизирован с SKUD')
                    else:
                        messages.warning(request, f'PINFL обновлен, но не синхронизирован с SKUD: {sync_result.get("error_details", "Неизвестная ошибка")}')
                    
                    return redirect('employee_detail', employee_id=employee.id)
                except Exception as e:
                    messages.error(request, f'Ошибка при обновлении PINFL: {str(e)}')
        else:
            form = PINFLUpdateForm(employee=employee)
        
        context = {
            'form': form,
            'employee': employee,
        }
        
        return render(request, 'employees/update_pinfl.html', context)
        
    except Employee.DoesNotExist:
        messages.error(request, 'Сотрудник не найден')
        return redirect('employees_list')


@require_login
def get_departments(request):
    """AJAX endpoint для получения отделов по организации"""
    organization_id = request.GET.get('organization_id')
    
    if organization_id:
        departments = Department.objects.filter(organization_id=organization_id).values('id', 'name')
        return JsonResponse(list(departments), safe=False)
    
    return JsonResponse([], safe=False)


@require_login
def get_divisions(request):
    """AJAX endpoint для получения подразделений по отделу"""
    department_id = request.GET.get('department_id')
    
    if department_id:
        divisions = Division.objects.filter(department_id=department_id).values('id', 'name')
        return JsonResponse(list(divisions), safe=False)
    
    return JsonResponse([], safe=False)