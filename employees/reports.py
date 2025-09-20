"""
Модуль для генерации отчётов системы учёта рабочего времени
"""

import io
import csv
from datetime import datetime, date, timedelta
from decimal import Decimal
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Q, Sum, Count, Avg
from django.contrib.auth.models import User

from .models import (
    Employee, WorkSession, WorkDaySummary, WorkTimeAuditLog,
    SKUDDevice, SKUDEvent
)


class WorkTimeReportGenerator:
    """Генератор отчётов по рабочему времени"""
    
    def __init__(self):
        self.date_format = "%d.%m.%Y"
        self.datetime_format = "%d.%m.%Y %H:%M"
    
    def generate_monthly_report_csv(
        self, 
        year: int, 
        month: int, 
        department_id: Optional[str] = None,
        employee_id: Optional[str] = None
    ) -> HttpResponse:
        """Генерация месячного отчёта в формате CSV"""
        
        # Определяем период
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # Получаем данные
        summaries = self._get_summaries_for_period(start_date, end_date, department_id, employee_id)
        
        # Создаём CSV ответ
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="worktime_report_{year}_{month:02d}.csv"'
        
        # Добавляем BOM для корректного отображения в Excel
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Заголовок
        writer.writerow([
            'ФИО', 'Табельный номер', 'Отдел', 'Подразделение',
            'Дата', 'Статус дня', 'Первый вход', 'Последний выход',
            'Отработано (часы)', 'Ожидаемо (часы)', 'Переработка (часы)', 'Недоработка (часы)',
            'Количество сессий', 'Проблемы'
        ])
        
        # Данные
        for summary in summaries:
            writer.writerow([
                summary.employee.full_name,
                summary.employee.employee_id,
                summary.employee.department.name if summary.employee.department else '',
                summary.employee.division.name if summary.employee.division else '',
                summary.date.strftime(self.date_format),
                summary.get_status_display(),
                summary.first_entry.strftime(self.datetime_format) if summary.first_entry else '',
                summary.last_exit.strftime(self.datetime_format) if summary.last_exit else '',
                round(summary.total_hours, 2),
                round(summary.expected_hours, 2),
                round(summary.overtime_hours, 2),
                round(summary.underwork_hours, 2),
                summary.sessions_count,
                'Да' if (summary.has_missing_exit or summary.has_manual_corrections) else 'Нет'
            ])
        
        return response
    
    def generate_monthly_report_xlsx(
        self, 
        year: int, 
        month: int, 
        department_id: Optional[str] = None,
        employee_id: Optional[str] = None
    ) -> HttpResponse:
        """Генерация месячного отчёта в формате Excel"""
        
        # Определяем период
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        # Получаем данные
        summaries = self._get_summaries_for_period(start_date, end_date, department_id, employee_id)
        
        # Создаём Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Отчёт {month:02d}.{year}"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = [
            'ФИО', 'Табельный номер', 'Отдел', 'Подразделение',
            'Дата', 'Статус дня', 'Первый вход', 'Последний выход',
            'Отработано (часы)', 'Ожидаемо (часы)', 'Переработка (часы)', 'Недоработка (часы)',
            'Количество сессий', 'Проблемы'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Данные
        for row, summary in enumerate(summaries, 2):
            data = [
                summary.employee.full_name,
                summary.employee.employee_id,
                summary.employee.department.name if summary.employee.department else '',
                summary.employee.division.name if summary.employee.division else '',
                summary.date.strftime(self.date_format),
                summary.get_status_display(),
                summary.first_entry.strftime(self.datetime_format) if summary.first_entry else '',
                summary.last_exit.strftime(self.datetime_format) if summary.last_exit else '',
                round(summary.total_hours, 2),
                round(summary.expected_hours, 2),
                round(summary.overtime_hours, 2),
                round(summary.underwork_hours, 2),
                summary.sessions_count,
                'Да' if (summary.has_missing_exit or summary.has_manual_corrections) else 'Нет'
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                # Выделяем проблемные строки
                if summary.has_missing_exit or summary.has_manual_corrections:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Автоширина колонок
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, ws.max_row + 1):
                cell_value = ws[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Создаём ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="worktime_report_{year}_{month:02d}.xlsx"'
        
        wb.save(response)
        return response
    
    def generate_employee_detailed_report(
        self, 
        employee_id: str, 
        start_date: date, 
        end_date: date
    ) -> HttpResponse:
        """Генерация детального отчёта по сотруднику"""
        
        try:
            employee = Employee.objects.get(id=employee_id)
        except Employee.DoesNotExist:
            return HttpResponse("Сотрудник не найден", status=404)
        
        # Получаем данные
        summaries = WorkDaySummary.objects.filter(
            employee=employee,
            date__gte=start_date,
            date__lte=end_date
        ).order_by('date')
        
        sessions = WorkSession.objects.filter(
            employee=employee,
            date__gte=start_date,
            date__lte=end_date
        ).order_by('date', 'start_time')
        
        # Создаём Excel файл
        wb = openpyxl.Workbook()
        
        # Лист 1: Сводка по дням
        ws_summary = wb.active
        ws_summary.title = "Сводка по дням"
        
        # Лист 2: Детали сессий
        ws_sessions = wb.create_sheet("Детали сессий")
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Заполняем лист сводки
        summary_headers = [
            'Дата', 'Статус', 'Первый вход', 'Последний выход',
            'Отработано (ч)', 'Ожидаемо (ч)', 'Переработка (ч)', 'Недоработка (ч)',
            'Сессий', 'Проблемы'
        ]
        
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row, summary in enumerate(summaries, 2):
            data = [
                summary.date.strftime(self.date_format),
                summary.get_status_display(),
                summary.first_entry.strftime(self.datetime_format) if summary.first_entry else '',
                summary.last_exit.strftime(self.datetime_format) if summary.last_exit else '',
                round(summary.total_hours, 2),
                round(summary.expected_hours, 2),
                round(summary.overtime_hours, 2),
                round(summary.underwork_hours, 2),
                summary.sessions_count,
                'Да' if (summary.has_missing_exit or summary.has_manual_corrections) else 'Нет'
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_summary.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Заполняем лист сессий
        session_headers = [
            'Дата', 'Начало', 'Окончание', 'Длительность (ч)',
            'Статус', 'Причина корректировки'
        ]
        
        for col, header in enumerate(session_headers, 1):
            cell = ws_sessions.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row, session in enumerate(sessions, 2):
            data = [
                session.date.strftime(self.date_format),
                session.start_time.strftime(self.datetime_format),
                session.end_time.strftime(self.datetime_format) if session.end_time else '',
                round(session.duration_hours, 2),
                session.get_status_display(),
                session.manual_reason or ''
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_sessions.cell(row=row, column=col, value=value)
                cell.border = border
        
        # Автоширина колонок
        for ws in [ws_summary, ws_sessions]:
            for col in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for row in range(1, ws.max_row + 1):
                    cell_value = ws[f"{column_letter}{row}"].value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Создаём ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"employee_report_{employee.employee_id}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    def generate_department_statistics_report(
        self, 
        department_id: str, 
        start_date: date, 
        end_date: date
    ) -> HttpResponse:
        """Генерация отчёта по статистике отдела"""
        
        # Получаем сотрудников отдела
        employees = Employee.objects.filter(
            department_id=department_id,
            is_active=True
        )
        
        if not employees.exists():
            return HttpResponse("В отделе нет активных сотрудников", status=404)
        
        department_name = employees.first().department.name
        
        # Создаём Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Статистика отдела {department_name}"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = [
            'ФИО', 'Табельный номер', 'Должность', 'Ставка',
            'Дней в периоде', 'Дней присутствия', 'Дней отсутствия', 'Дней по уважительной причине',
            'Отработано (ч)', 'Ожидаемо (ч)', 'Переработка (ч)', 'Недоработка (ч)',
            'Эффективность (%)', 'Проблемных дней'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Данные по каждому сотруднику
        row = 2
        for employee in employees:
            summaries = WorkDaySummary.objects.filter(
                employee=employee,
                date__gte=start_date,
                date__lte=end_date
            )
            
            total_days = (end_date - start_date).days + 1
            present_days = summaries.filter(status='present').count()
            absent_days = summaries.filter(status='absent').count()
            excused_days = summaries.filter(status='excused').count()
            problem_days = summaries.filter(
                Q(has_missing_exit=True) | Q(has_manual_corrections=True)
            ).count()
            
            total_hours = sum(s.total_hours for s in summaries)
            expected_hours = sum(s.expected_hours for s in summaries)
            overtime_hours = sum(s.overtime_hours for s in summaries)
            underwork_hours = sum(s.underwork_hours for s in summaries)
            
            efficiency = (total_hours / expected_hours * 100) if expected_hours > 0 else 0
            
            data = [
                employee.full_name,
                employee.employee_id,
                employee.get_position_display(),
                f"{employee.work_fraction * 100:.0f}%",
                total_days,
                present_days,
                absent_days,
                excused_days,
                round(total_hours, 2),
                round(expected_hours, 2),
                round(overtime_hours, 2),
                round(underwork_hours, 2),
                round(efficiency, 1),
                problem_days
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                
                # Выделяем проблемных сотрудников
                if problem_days > 0 or efficiency < 80:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            row += 1
        
        # Автоширина колонок
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, ws.max_row + 1):
                cell_value = ws[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Создаём ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"department_stats_{department_name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    def _get_summaries_for_period(
        self, 
        start_date: date, 
        end_date: date,
        department_id: Optional[str] = None,
        employee_id: Optional[str] = None
    ):
        """Получение сводок за период с фильтрацией"""
        
        queryset = WorkDaySummary.objects.filter(
            date__gte=start_date,
            date__lte=end_date
        ).select_related('employee', 'employee__department', 'employee__division')
        
        if department_id:
            queryset = queryset.filter(employee__department_id=department_id)
        
        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)
        
        return queryset.order_by('employee__last_name', 'employee__first_name', 'date')
